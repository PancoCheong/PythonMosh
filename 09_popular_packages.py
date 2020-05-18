# 09_popular_packages.py
#
#
# wxPython - Graphic User Interface
# SQLAlchemy - SQL Toolkit and Object Relational Mapper
# SciPy - algorithms and mathematical tools for scientific calculations (contain numpy, pandas, matplotlib)
# pyGame - create Game by Python
# Matplotlib - plot the graph for data analysis
#
# Application Programming Interface API - endpoint of application for exchange data
# REST - a bunch of conventionns of rules and conventions that we have to follow to build or consume API for exchange data.
#
# gmail
#
# Request URL: https://www.yelp.com/developers/documentation/v3/business_search
# Request Method: GET
# Status Code: 200
# Remote Address: 104.16.57.23:443
# Referrer Policy: no-referrer-when-downgrade
#
## Response Headers ##
# cache-control: max-age=0, must-revalidate, no-cache, no-store, private
# cache-control: no-transform
# cf-cache-status: DYNAMIC
# cf-ray: 588bd55efa4aaff0-MFM
# cf-request-id: 024b43af5d0000aff0f3af3200000001
# content-encoding: gzip
#
#
### Google search: Yelp API ###
### Yelp Fusion --> Documentation ###
# https://www.yelp.com/developers/documentation/v3/business_search
#
# 1. Create App (Menu item under General on the left)
#     (VPN to US or HK to create and get the key)
#
# Business Endpoints --> Business Search
# https://www.yelp.com/developers/documentation/v3/business_search
# GET https://api.yelp.com/v3/businesses/search

import pandas as pd
import html
import pprint
import json
import requests
import random
import numpy as np
import openpyxl
import PyPDF2
from selenium import webdriver
from bs4 import BeautifulSoup
from twilio.rest import Client
import config
import requests     # HTTP requests
url = "https://api.yelp.com/v3/businesses/search"
# move to config.py
# #api_key = "dsfasdddddasfsf-afdsfafasffa"
headers = {
    "Authorization": "Bearer " + config.api_key
}
params = {
    "term": "Barber",
    "location": "NYC"
}
response = requests.get(url, headers=headers, params=params)
print(response)                     # http response code
print(response.status_code)         # http response code
print(response.headers)             # http headers
print(response.headers["Date"])     # http Date attribute in headers
# print(response.text)              # http response payload body
# print(type(response.text))        # output:<class str>
#
# output:<Response [400]>           #Bad Request
# output:<Response [401]>           #Bad Request
# output: {"businesses": [{"id": "KzRmCW_Fe7aW0qF3VjIOxg", "alias": "next-level-barber-shop-new-york", "name": "Next Level Barber Shop", "image_url": "https://s3-media3.fl.yelpcdn.com/bphoto/3or7-J9_FYrV-CInL30HEg/o.jpg", "is_closed": false, "url": "https://www.yelp.com/


# Introduction --> Authenticate --> authentication guide
# https://www.yelp.com/developers/documentation/v3/authentication
#
# Create an app to obtain your private API Key.
# Authenticate API calls with the API Key.
# -    set the "Authorization" HTTP header value as Bearer API_KEY.
# VPN US IP: 199.187.211.100
#
businesses = response.json()["businesses"]    # convert JSON to dictionary
# print(businesses)
#
# output: [{'id': 'KzRmCW_Fe7aW0qF3VjIOxg', 'alias': 'next-level-barber-shop-new-york', 'name': 'Next Level Barber Shop', 'image_url': 'https://s3-media3.fl.yelpcdn.com/bphoto/3or7-J9_FYrV-CInL30HEg/o.jpg', 'is_closed': False, 'url': 'https://www.yelp.com/
#
#
for business in businesses:
    print(business["rating"], " - ", business["name"])
#
# use list comprehension
# [item for item in list if condition]
# business with rating > 4.5
names = [business["name"]
         for business in businesses if business["rating"] > 4.5]
print(names)
#
#

### Open Trivia Database - for API testing ###
# API --> Number of Questions = 1 --> Category = Entertainment: Music, Difficulty = easy, type = Mutliple choice --> generate API URL
# https://opentdb.com/api.php?amount=1&category=12&difficulty=easy&type=multiple
#
#import requests
r = requests.get(
    "https://opentdb.com/api.php?amount=1&category=12&difficulty=easy&type=multiple")
print(r.status_code)    # output:200
# output:{"response_code":0,"results":[{"category":"Entertainment: Music","type":"multiple","difficulty":"easy","question":"What Led Zeppelin album contains &quot;Stairway to Heaven&quot;?","correct_answer":"Led Zeppelin IV","incorrect_answers":["Houses of the Holy","Physical Graffiti","Led Zeppelin III"]}]}
print(r.text)
print(type(r.text))     # output:<class 'str'>
#
#
#import json
# convert json str to dict
question = json.loads(r.text)
# output:{'response_code': 0, 'results': [{'category': 'Entertainment: Music', 'type': 'multiple', 'difficulty': 'easy', 'question': 'What Led Zeppelin album contains &quot;Stairway to Heaven&quot;?', 'correct_answer': 'Led Zeppelin IV', 'incorrect_answers': ['Houses of the Holy', 'Physical Graffiti', 'Led Zeppelin III']}]}
print(question)
print(type(question))   # output:<class 'dict'>
#
# print json dict in human easy-to-read format
#import pprint
# output:
# {'response_code': 0,
# 'results': [{'category': 'Entertainment: Music',
#              'correct_answer': 'Led Zeppelin IV',
#              'difficulty': 'easy',
#              'incorrect_answers': ['Houses of the Holy',
#                                    'Physical Graffiti',
#                                    'Led Zeppelin III'],
#              'question': 'What Led Zeppelin album contains &quot;Stairway to '
#                          'Heaven&quot;?',
#              'type': 'multiple'}]}
pprint.pprint(question)
#
# extrac information from json dict
# output:Entertainment: Music
print(question['results'][0]['category'])
# output:['Doin&#039; It Right', 'Instant Crush', 'The Game of Love']
# print value of the key=incorrect_answers which is a list of questions
print(question['results'][0]['incorrect_answers'])
#
print(type(data['results']))                            # output:<class 'list'>
print(type(data['results'][0]))                         # output:<class 'dict'>
print(type(data['results'][0]['incorrect_answers']))    # output:<class 'list'>
#
person = {'Name': 'Panco', 'Height': 1.82}
person_json = json.dumps(person)
print(person_json)          # output:{"Name": "Panco", "Height": 1.82}
print(type(person_json))    # output:<class 'str'>


### Exercise - Create a quizzing game ###
# Make an HTTP request to the Open Trivia API at each round of the game to get a new question
# an present it to the user to answer.
# show the answer if it is correct or not at each round
# keep the user's score
# At the end of each round ask the user if he wants to play again.
# Keep playing forever until the user types "quit".

#import requests
#import json
#import pprint
#import random
#import html
#
# keep the score
score_correct = 0
score_incorrect = 0

url = "https://opentdb.com/api.php?amount=1"
endGame = ""

while endGame != "quit":
    r = requests.get(url)
    if (r.status_code != 200):
        endGame = input(
            "Sorry, there was a problem retrieving the question. Press enter to try again or type 'quit' to quit  the game.")
    else:
        valid_answer = False
        answer_number = 1
        data = json.loads(r.text)
        question = data['results'][0]['question']
        answers = data['results'][0]['incorrect_answers']
        correct_answer = data['results'][0]['correct_answer']
        print(type(answers))
        # append the correct answer into list
        answers.append(correct_answer)
        random.shuffle(answers)             # shuffle the answers

        print(html.unescape(question) + "\n")

        # let user to input number instead of typing the answer in string
        # convert special char to readable char ie. &#039; -> '   &quot; -> "
        # https://dev.w3.org/html5/html-author/charref
        # https://www.tutorialspoint.com/html/html_ascii_codes.htm
        #
        for answer in answers:
            print(str(answer_number) + "- " + html.unescape(answer))
            answer_number += 1

        # input validation - input number and within the size of the list
        while valid_answer == False:
            user_answer = input("\nType the number of the correct answer: ")
            try:
                user_answer = int(user_answer)
                if user_answer > len(answers) or user_answer <= 0:
                    print("Invalid Answer")
                else:
                    valid_answer = True
            except:
                print("Invalid answer. Use only numbers.")

        # list index start from zero
        user_answer = answers[int(user_answer)-1]

        # keep track of score
        if user_answer == correct_answer:
            print("\nCongratulations! You answered correctly! The correct answer was: " +
                  html.unescape(correct_answer))
            score_correct += 1
        else:
            print("Sorry, " + html.unescape(user_answer) +
                  " is incorrect. The correct answer is: " + html.unescape(correct_answer))
            score_incorrect += 1

        # print the current result
        print("\n##########################")
        print("Your score is:")
        print("Correct answers: " + str(score_correct))
        print("Incorrect answers: " + str(score_incorrect))
        print("##########################")

        endGame = input(
            "\nPress enter to play again or type 'quit' to quit the game.").lower()

print("\nThanks for playing")


### hiding API keys ###
## don't store API key in source code ##
# put the key in separate file and exclude it from the git #
# ie. move api_key to config.py
#
#
# import config
# reference the variable
# config.api_key
# put the name "config.py" in .gitignore file
#
#
### use twilio to send SMS Text messaging ###
# support Voice, SMS, Video call and WhatsApp messaging #
# 1. use http requests to call API
# 2. use library in pip or pipenve, a wrapper around the API
# and it encapsulate all that http communcation,
# so you don't need to handle when to use GET, POST, PUT, DELETE etc
#
# sign up account: https://www.twilio.com/try-twilio
# hotmail
# nism@sky1ine<CAR3chars>
#
# Phone Numbers --> Get Started --> Get your first Twilio phone number
# +1 224 442 9684
#
# in terminal
# pipenv install twilio
#
#from twilio.rest import Client
#
# move to config.py
# account_sid = "adsfdfsafsa"
# auth_token = "aafdsfadsf"
client = Client(config.account_sid, config.auth_token)
# client.calls
# client.voice
# client.fax
# client.chat
#
# from_ has underscore
#
# call = client.messages.create(
#     to="+85366801224",
#     from_="+12244429684",
#     body="This message is sent from Python application from my Laptop by using Twilio"
# )

# print("created:", call.date_created, " sent:",
#       call.date_sent, " update:", call.date_updated)
#
#
### web scraping ###
# extract the data from HTML webpage #
#
# extract questions from stackoverflow
# https://stackoverflow.com/questions
#
# package to extract info from HTML or XML
# pipenv install beautifulsoup4
#
# HTTP requests
# pipenv install requests
#
# import requests
#from bs4 import BeautifulSoup
response = requests.get("https://stackoverflow.com/questions")
soup = BeautifulSoup(response.text, "html.parser")
#
# google chrome, inspect the question
#
#  <div id="questions" ...
#      <div class="question-summary" ...
#          <div class="summary" ...
#               <h3>
#                   <a href="...." class="question-hyperlink"> QUESTION </a>
#
# select the CSS element:   . class    # id
questions = soup.select(".question-summary")
print(type(questions[0]))               # output:<class 'bs4.element.Tag'>
# output:{'class': ['question-summary'], 'id': 'question-summary-61403882'}
print(questions[0].attrs)               # <class 'dict'>
#
# not recommended, throw exception if the attribute is not found
# print(questions[0]["id"])               # output:question-summary-61403938
# safer way to use get() and supply the default value if not found
print(questions[0].get("id", 0))        # output:question-summary-61403938
#
# output:[<a class="question-hyperlink" href="/questions/61404033/python-property-in-self-defined-object-can-not-update-using-self-setattr">Python property in self defined object can not update, using self.__setattr__</a>]
print(questions[0].select(".question-hyperlink"))
#
# select one only
# output:<a class="question-hyperlink" href="/questions/61404033/python-property-in-self-defined-object-can-not-update-using-self-setattr">Python property in self defined object can not update, using self.__setattr__</a>
print(questions[0].select_one(".question-hyperlink"))
#
# get the text inside <tag>text</tag>
# output:Python property in self defined object can not update, using self.__setattr__
print(questions[0].select_one(".question-hyperlink").getText())
#
print("----- questions -----")
for question in questions:
    print(question.select_one(".question-hyperlink").getText())
    print(question.select_one(".vote-count-post").getText())
#
#
### Browser Automation ###
## selenium ##
# pipenv install selenium
#
# download the driver for each browser (manually download)
# https://pypi.org/project/selenium/
#
# scroll down to Drivers
# Chrome:	https://sites.google.com/a/chromium.org/chromedriver/downloads
# Edge:	https://developer.microsoft.com/en-us/microsoft-edge/tools/webdriver/
# Firefox:	https://github.com/mozilla/geckodriver/releases
# Safari:	https://webkit.org/blog/6900/webdriver-support-in-safari-10/
#
# extract the chromedriver.exe to local folder in the PATH
# Windows: C:\Windows\system32
# MacOSX: /usr/local/bin
# from selenium import webdriver

# test the github login process
browser = webdriver.Chrome()
browser.get("https://github.com")
# find the HTML element by class, id, name and
#
# <a href="/login" class="HeaderMenu-link no-underline mr-3"
# data-hydro-click="{&quot;event_type&quot;:&quot;authentication.click&quot;,&quot;payload&quot;
# :{&quot;location_in_page&quot;:&quot;site header menu&quot;,&quot;repository_id&quot;
# :null,&quot;auth_type&quot;:&quot;SIGN_UP&quot;,&quot;originating_url&quot;
# :&quot;https://github.com/&quot;,&quot;user_id&quot;:null}}"
# data-hydro-click-hmac="cd4f672ed9a2fa51ea92c28de162e81edb2d11a2aad6884ec89a6d60b21b1cfb"
# data-ga-click="(Logged out) Header, clicked Sign in, text:sign-in">
#          Sign&nbsp;in
# </a>
#
### class is not unique ###
### no id ###
### no name ###
### can only use text ###
#
signin_link = browser.find_element_by_link_text("Sign in")    # case sensitive
signin_link.click()

#
# <input type="text" name="login" id="login_field"
# class="form-control input-block" tabindex="1" autocapitalize="off"
# autocorrect="off" autocomplete="username" autofocus="autofocus">
#
### id is unique - login_field ###
#
# <input type="password" name="password" id="password"
# class="form-control form-control input-block" tabindex="2"
# autocomplete="current-password">
#
### id is unique - password ###
#
username_box = browser.find_element_by_id("login_field")
username_box.send_keys("pancocheong")
password_box = browser.find_element_by_id("password")
password_box.send_keys("Abcd1234")
password_box.submit()
#
# validate if the page is correct
# do assertion
#
# throw exception if the content is not found
# the word PancoCheong is in the HTML source page
# assert "PancoCheongXXX" in browser.page_source
# output:assert "PancoCheongXXX" in browser.page_source
#        AssertionError
assert "PancoCheong" in browser.page_source
print("Use generic way to assert the function")
#
#
# Inspect the drop down menu
#
# <a role="menuitem" class="no-underline user-profile-link px-3 pt-2 pb-2 mb-n2 mt-n1 d-block"
# href="/PancoCheong" data-ga-click="Header, go to profile, text:Signed in as">
# Signed in as
# <strong class="css-truncate-target">PancoCheong</strong>
# </a>
#
### class has unique - user-profile-link ###
#
# use specific way to do assertion
profile_link = browser.find_element_by_class_name("user-profile-link")
# output:profile_link: <selenium.webdriver.remote.webelement.WebElement (session="bcdd2279ca8d668079729a48fb0ac95c", element="f68c13d9-da3a-47db-989a-f5a003b95e45")>
print("profile_link:", profile_link)
link_label = profile_link.get_attribute(
    "innerHTML")  # extract the Text inside <HTML Tag>
# output:link_label Signed in as <strong class="css-truncate-target">PancoCheong</strong>
print("link_label:", link_label)
assert "PancoCheong" in link_label
print("use specific way to assert the function")
#
# google search: Python Selenium
# https://selenium-python.readthedocs.io/
# recommended to read: Waits and Page Objects
#
# close the browser
browser.quit()
#
#
#
### working with PDF ###
# pipenv install pypdf2
#
# import PyPDF2
with open("first.pdf", "rb") as file:   # read and binary mode
    # pass in File stream (in binary mode)
    reader = PyPDF2.PdfFileReader(file)
    print(reader.numPages)              # output:1
    page = reader.getPage(0)            # page index starts 0
    # extract text (output text is a mess, not working)
    print(page.extractText())
    # below change is in memory
    # page.rotateClockwise(90)              # rotate angle
    # page.rotateCounterClockwise(90)     # rotate angle
    # page.scale(1.5, 1.5)                # scale by x, y factor
    # page.scaleBy(1.5)                   # scale by single factor
    # page.scaleTo(100, 200)              # scale to width, height
#
    # add watermark to page
    with open("watermark.pdf", "rb") as watermarkfile:
        watermark = PyPDF2.PdfFileReader(watermarkfile)
        page.mergePage(watermark.getPage(0))

    #
        writer = PyPDF2.PdfFileWriter()     #
        writer.addPage(page)                # append to the end of file
        # writer.insertBlankPage(0)           # index = 0 (1st page)
        # writer.insertPage(page, 0)  # index = 0 (1st page)
    #
    #   write change to file
        with open("approved.pdf", "wb") as output:   # write, binary
            writer.write(output)

### rotate all pages ###
# for pagenum in range(pdfReader.numPages):
#         page = pdfReader.getPage(pagenum)
#         page.rotateClockwise(rotation)
#         pdfWriter.addPage(page)
#
# alternative PDF package:
# https://github.com/jalan/pdftotext
#
### merge 2 PDF files ###
merger = PyPDF2.PdfFileMerger()
files_names = ["first.pdf", "second.pdf"]
for file_name in files_names:
    merger.append(file_name)
merger.write("combined.pdf")
#
#
#
### Working with Excel ###
# https://openpyxl.readthedocs.io/en/stable/
#
# pipenv install openpyxl
#
#import openpyxl

# either create an empty workbook
# wb = openpyxl.Workbook()
#
# or load the existing workbook
wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)        # output:['Sheet1']
sheet = wb["Sheet1"]        # case sensitive
#
#
sheetnames = wb.sheetnames
ws = wb[sheetnames[0]]      # load first worksheet
print(ws)                   # output:<Worksheet "Sheet1">
ws = wb.active              # load active worksheet
print(ws)                   # output:<Worksheet "Sheet1">

# copy workbook
source = wb.active
target = wb.copy_worksheet(source)
#
# create new worksheet
# index 0 - put it before existing worksheet
ws1 = wb.create_sheet("MyNewSheet", 0)
ws2 = wb.create_sheet("Sheet2")           # put it after existing worksheets
ws3 = wb.create_sheet("Penultimate", -1)       # put it on 2nd last
# remove worksheet
# wb.remove_sheet("Sheet2")
wb.remove(ws3)      # use object, not sheet name
#
# change sheet tab color (RRGGBB)
ws1.sheet_properties.tabColor = "FFA500"
#
cell = sheet["a1"]
print("value:", cell.value)         # output:value:transaction_id
# change cell value
# cell.value = "new value"
#
print("row:", cell.row)                  # output:row:1
print("column:", cell.column)            # output:column:A
print("coordinate:", cell.coordinate)    # output:coordinate:A1
#
# other way to access the cell - easier to iterate
# cell = sheet.cell(1,1)                 # row and column numbers
cell = sheet.cell(row=1, column=1)       # same as sheet["a1"]
#
print("max_row:", sheet.max_row)         # output:max_row:4
print("max_column:", sheet.max_column)   # output:max_column:3
#
# range() - index start by 0 by default (if single parameter)
# last index is excluded, add 1 to access it
for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, column)
        # output:transaction_id product_id .... 1003 3 7.95
        print(cell.value)
#
# or use iter_rows()
for row in sheet.iter_rows(min_row=1, max_row=3,
                           min_col=1, max_col=2):
    for cell in row:
        # output:<Cell 'Sheet1'.A1> <Cell 'Sheet1'.B1> ... <Cell 'Sheet1'.A3> <Cell 'Sheet1'.B3>
        print(cell)

# all the cell in a column (tuple)
column_a = sheet["a"]       # tuple
# output:(<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.A2>, <Cell 'Sheet1'.A3>, <Cell 'Sheet1'.A4>)
print(column_a)
cells = sheet["a:c"]        # tuple of tuple, each tuple is 1 column
# output:((<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.A2>, <Cell 'Sheet1'.A3>, <Cell 'Sheet1'.A4>), (<Cell 'Sheet1'.B1>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.B4>), ( ....
print(cells)
cell_range = sheet["a1:c3"]  # tuple of tuple, each tuple is 1 column
# output:((<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>), (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>),
print(cell_range)

# all the cells in row 1
sheet[1]
# all the cells in 1st 3 rows
sheet[1:3]

# append a row at the end, parameter=list of values
sheet.append([1, 2, 3])
sheet.insert_rows(1)        # insert 1 row at 1st row
sheet.insert_cols(2, 3)      # at 2 column, insert 4 columns
sheet.delete_rows(4)        # delete row 4
sheet.delete_cols(3, 1)      # at 3 column, delete 1 column
#
# print all title
for sheet in wb:
    # output:MyNewSheet Sheet1 Sheet1 Copy Sheet2
    print(sheet.title, end=" ")
print("")
# save it as new file
wb.save("transactions2.xlsx")
#
#
### Command Query Separation Principle ###
# query command should never update the state of object #
#
wb = openpyxl.load_workbook("transactions.xlsx")
sheet = wb["Sheet1"]        # case sensitive
for row in range(1, 10):
    cell = sheet.cell(row, 1)
    # output:transaction_id 1001 1002 1003 None None None None None
    print(cell.value)
sheet.append([1, 2, 3])
wb.save("transactions3.xlsx")
#
# 5 empty rows are added before the appened rows - non-expected behavior
#
#
### NumPy ###
# just like generic type of list - store only single data type
# process faster and use less memory
# most frequent used package for data analytic
#
# pipenv install numpy
#
# import numpy as np
# 1-D array
array = np.array([1, 2, 3])
print(array)
print(type(array))
# 2-D array - aka. matrix
array_2d = np.array([[1, 2, 3], [4, 5, 6]])     # 2 rows and 3 columns
print(array_2d)
print(array_2d.shape)                           # output:(2, 3)

# array with all zero value (floating number by default)
array = np.zeros((3, 4))                # shape = 3 rows, 4 columns
print(array)
#
array = np.zeros((3, 4), dtype=int)     # change data type
print(array)
#
# array with all value of 1
array = np.ones((3, 4))                 # shape = 3 rows, 4 columns
print(array)

#
# array with all value of specified number 99
array = np.full((3, 4), 99, dtype=int)   # shape = 3 rows, 4 columns
print(array)
#
# import random
#
# array with random value between 0 and 1
array = np.random.random((3, 4))
print(array)
#
# access the matrix
print(array[0, 0])
#
# conditional check on each item
print(array > 0.4)      # output: bool
#
# filter the output (only > 0.4)
print(array[array > 0.4])
#
# sum of all items
print(np.sum(array))
# math functions
print(np.floor(array))
print(np.ceil(array))
print(np.round(array))
#
first = np.array([[1, 2, 3], [4, 5, 6]])
second = np.array([[11, 12, 13], [14, 15, 16]])
print(first + second)   # add position by position (ie. 1+11, 2+12)
print(first * 3)        # multiple each item
#
# eg. convert inch to cm
numbers_inch = np.array([1, 2, 3])
numbers_cm = numbers_inch * 2.54
print(numbers_cm)
#
# use standard Python list comprehension
numbers_inch = [1, 2, 3]
numbers_cm = [number * 2.54 for number in numbers_inch]
print(numbers_cm)
#
#
#
### Pandas and xlrd ###
# pip install pandas
# for Excel
# pip install xlrd
#import pandas as pd
file = pd.ExcelFile("sales.xlsx")
print(file.sheet_names)                 # output:['sales', 'customers']
sheet = file.parse('sales')
#
## print the worksheet ##
# output:
#         Date             Customer  Invoice  Amount
# 0 2018-12-01  Steel Brothers Inc.       98    1340
# 1 2018-12-10             MMC Inc.       99    1900
# 2 2018-12-12             MMC Inc.      100    2900
# 3 2018-12-18  Steel Brothers Inc.      101     977
# 4 2018-12-21     Steel & Iron LLC      102    3400
print(sheet)
#
print(type(sheet))  # output:<class 'pandas.core.frame.DataFrame'>
#
## print the Date column ##
# output:
# 0   2018-12-01
# 1   2018-12-10
# 2   2018-12-12
# 3   2018-12-18
# 4   2018-12-21
# Name: Date, dtype: datetime64[ns]
print(sheet.Date)
#
# calculate the SUM of Amount column
print("Sum of Amount =", sheet.Amount.sum())   # output: Sum of Amount = 10517
#
## Get the 1st data row from Excel ##
# output:
# Date        2018-12-01 00:00:00
# Customer    Steel Brothers Inc.
# Invoice                      98
# Amount                     1340
# Name: 0, dtype: object
print(sheet.loc[0])
#
# print the value of Amount in 1st data row
print(sheet.loc[0, "Amount"])   # output:1340
#
#
## search the sales made by a specified customer ##
sheet.set_index("Customer", inplace=True)
# instead of using index number to fetch the rows
# use the value in Customer column to search
# output:
#               Date  Invoice  Amount
# Customer
# MMC Inc. 2018-12-10       99    1900
# MMC Inc. 2018-12-12      100    2900
print(sheet.loc["MMC Inc."])
sheet = sheet.reset_index()
#
# print Invoice column
# output:
# 0     98
# 1     99
# 2    100
# 3    101
# 4    102
# Name: Invoice, dtype: int64
print(sheet["Invoice"])
print(type(sheet["Invoice"]))   # output:<class 'pandas.core.series.Series'>
#
# search invoice number = 99
# output:
#   Customer       Date  Invoice  Amount
# 1  MMC Inc. 2018-12-10       99    1900
print(sheet.loc[sheet["Invoice"] == 99])
#
# search the rows that has amount > 2000
# output:
#           Customer       Date  Invoice  Amount
# 2          MMC Inc. 2018-12-12      100    2900
# 4  Steel & Iron LLC 2018-12-21      102    3400
print(sheet.loc[sheet["Amount"] > 2000])
#
# search the row which has highest amount
# output:
# Customer       Steel & Iron LLC
# Date        2018-12-21 00:00:00
# Invoice                     102
# Amount                     3400
# Name: 4, dtype: object
print(sheet.loc[sheet["Amount"].idxmax()])
#
# search the customer name which has highest amount
print(sheet.loc[sheet["Amount"].idxmax()]
      ["Customer"])  # output:Steel & Iron LLC
#
# search the rows which has amount > 1800
# output:
#           Customer       Date  Invoice  Amount
# 1          MMC Inc. 2018-12-10       99    1900
# 2          MMC Inc. 2018-12-12      100    2900
# 4  Steel & Iron LLC 2018-12-21      102    3400
print(sheet.loc[sheet["Amount"] > 1800])
#
# search the customer name who has amount > 1800
# output:
# 1            MMC Inc.
# 2            MMC Inc.
# 4    Steel & Iron LLC
# Name: Customer, dtype: object
print(sheet.loc[sheet["Amount"] > 1800]["Customer"])
#
# search the distinct customer name who has amount > 1800
# output: ['MMC Inc.' 'Steel & Iron LLC']
print(sheet.loc[sheet["Amount"] > 1800]["Customer"].unique())
#
# search the 1st distinct customer name who has amount > 1800
# output:MMC Inc.
print(sheet.loc[sheet["Amount"] > 1800]["Customer"].unique()[0])
#
# loop thru the customer name
# output:
# MMC Inc.
# Steel & Iron LLC
for customer in sheet.loc[sheet["Amount"] > 1800]["Customer"].unique():
    print(customer)
